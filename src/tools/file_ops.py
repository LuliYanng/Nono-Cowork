"""
File operation tools — read, write, and edit files.

Supports binary formats: PDF (.pdf), Excel (.xlsx/.xls), Word (.docx).
Binary files are auto-converted to text/Markdown when read.
"""

import os
import shutil
from datetime import datetime
from tools.registry import tool


# ————— Binary format converters —————

def _read_pdf(path: str) -> str:
    """Extract text from a PDF file. Falls back to OCR for scanned pages."""
    import pymupdf
    doc = pymupdf.open(path)
    pages = []
    ocr_used = False

    for i, page in enumerate(doc, 1):
        # Try normal text extraction first
        text = page.get_text().strip()
        if text:
            pages.append(f"--- Page {i} ---\n{text}")
            continue

        # No text — try OCR (for scanned pages)
        try:
            tp = page.get_textpage_ocr(language="chi_sim+eng", dpi=300)
            text = page.get_text("text", textpage=tp).strip()
            if text:
                pages.append(f"--- Page {i} (OCR) ---\n{text}")
                ocr_used = True
        except Exception:
            pass  # OCR not available or failed, skip this page

    doc.close()
    if not pages:
        return f"📄 {path} (PDF, {doc.page_count} pages)\n(No extractable text, and OCR found nothing. This PDF may contain only non-text images.)"
    suffix = " [OCR used for some pages]" if ocr_used else ""
    return f"📄 {path} (PDF, {len(pages)} pages with text{suffix})\n\n" + "\n\n".join(pages)


def _read_excel(path: str) -> str:
    """Read Excel file and return as Markdown tables."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        if not rows:
            continue

        # Build Markdown table
        # Use first row as header
        header = rows[0]
        col_count = len(header)
        header_strs = [str(c) if c is not None else "" for c in header]
        lines = [
            "| " + " | ".join(header_strs) + " |",
            "| " + " | ".join(["---"] * col_count) + " |",
        ]
        for row in rows[1:]:
            # Pad row if shorter than header
            cells = list(row) + [None] * (col_count - len(row))
            lines.append("| " + " | ".join(str(c) if c is not None else "" for c in cells[:col_count]) + " |")

        sheet_text = f"### Sheet: {name} ({len(rows)-1} data rows)\n\n" + "\n".join(lines)
        sheets.append(sheet_text)
    wb.close()

    if not sheets:
        return f"📄 {path} (Excel, {len(wb.sheetnames)} sheets)\n(All sheets are empty.)"
    return f"📄 {path} (Excel)\n\n" + "\n\n".join(sheets)


def _read_docx(path: str) -> str:
    """Extract text from a Word document."""
    from docx import Document
    doc = Document(path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return f"📄 {path} (Word)\n(No text content found.)"
    return f"📄 {path} (Word, {len(paragraphs)} paragraphs)\n\n" + "\n".join(paragraphs)


# Extension → converter mapping
_BINARY_READERS: dict[str, callable] = {
    ".pdf": _read_pdf,
    ".xlsx": _read_excel,
    ".xls": _read_excel,
    ".docx": _read_docx,
}


# ————— Snapshot: auto-backup before Agent edits —————
SNAPSHOT_DIR = ".agent_snapshots"
MAX_SNAPSHOTS = 50  # Keep at most this many snapshot folders


def _get_workspace():
    """Get the workspace root (same logic as prompt.py)."""
    ws = os.getenv("WORKSPACE_DIR", "").strip()
    if ws:
        return os.path.expanduser(ws)
    try:
        from tools.syncthing import SyncthingClient
        st = SyncthingClient()
        folders = st.get_folders()
        if folders:
            return folders[0]["path"]
    except Exception:
        pass
    return os.path.expanduser("~/")


_stignore_checked = False

# Essential patterns that must be in .stignore for a clean Agent workspace sync.
# Uses (?d) prefix = Syncthing will delete local copies of newly-ignored files on remote.
_REQUIRED_STIGNORE_PATTERNS = [
    ("Python venvs", [
        "(?d)**/*venv*",
        "(?d)**/env",
    ]),
    ("Python caches", [
        "(?d)**/__pycache__",
        "(?d)**/*.pyc",
        "(?d)**/*.pyo",
        "(?d)**/*.egg-info",
    ]),
    ("Node.js", [
        "(?d)**/node_modules",
    ]),
    ("IDE and system files", [
        "(?d)**/.idea",
        "(?d)**/.vscode",
        "(?d)**/.DS_Store",
    ]),
    ("Git repos", [
        "(?d)**/.git",
    ]),
    ("Agent internals", [
        "(?d).agent_snapshots",
        "(?d).stversions",
    ]),
]


def _ensure_stignore(workspace: str):
    """Ensure .stignore contains all essential patterns for a clean Agent workspace sync.

    Checks each required pattern and appends any missing ones.
    Idempotent — safe to call multiple times.
    """
    global _stignore_checked
    if _stignore_checked:
        return
    _stignore_checked = True

    try:
        stignore_path = os.path.join(workspace, ".stignore")

        # Read existing content
        existing = ""
        if os.path.exists(stignore_path):
            with open(stignore_path, "r") as f:
                existing = f.read()

        # Collect missing patterns
        missing_sections = []
        for section_name, patterns in _REQUIRED_STIGNORE_PATTERNS:
            missing = [p for p in patterns if p not in existing]
            if missing:
                missing_sections.append((section_name, missing))

        if not missing_sections:
            return  # All patterns already present

        # Append missing patterns
        with open(stignore_path, "a") as f:
            f.write("\n// ── Auto-managed by Agent (do not remove) ──\n")
            for section_name, patterns in missing_sections:
                f.write(f"// {section_name}\n")
                for p in patterns:
                    f.write(f"{p}\n")
    except Exception:
        pass  # Non-critical — don't break agent startup


def _snapshot_file(file_path: str) -> str | None:
    """Save a backup copy of a file before modifying it.

    Returns the snapshot path, or None if snapshot failed (non-fatal).
    """
    try:
        workspace = _get_workspace()
        snap_base = os.path.join(workspace, SNAPSHOT_DIR)

        # Ensure .agent_snapshots is in .stignore (only check once)
        _ensure_stignore(workspace)

        # Build snapshot path with microseconds to avoid same-second collision
        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        try:
            rel = os.path.relpath(file_path, workspace)
        except ValueError:
            rel = os.path.basename(file_path)

        snap_path = os.path.join(snap_base, ts, rel)
        os.makedirs(os.path.dirname(snap_path), exist_ok=True)
        shutil.copy2(file_path, snap_path)

        # Cleanup old snapshots if too many
        try:
            snap_dirs = sorted(os.listdir(snap_base))
            # Filter out non-directories
            snap_dirs = [d for d in snap_dirs if os.path.isdir(os.path.join(snap_base, d))]
            while len(snap_dirs) > MAX_SNAPSHOTS:
                oldest = snap_dirs.pop(0)
                shutil.rmtree(os.path.join(snap_base, oldest), ignore_errors=True)
        except Exception:
            pass

        return snap_path
    except Exception:
        return None  # Snapshot failure should never block the edit


@tool(
    name="read_file",
    description="Read file contents with optional line range. Supports text files, PDF, Excel (.xlsx), and Word (.docx) — binary formats are automatically converted to text. Use this to view code, configs, documents, spreadsheets, etc.",
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "File path (absolute or relative).",
            },
            "start_line": {
                "type": "integer",
                "description": "Starting line number (1-indexed). If not specified, starts from the beginning.",
            },
            "end_line": {
                "type": "integer",
                "description": "Ending line number (inclusive). If not specified, reads to the end of the file.",
            },
        },
        "required": ["path"],
    },
)
def read_file(path: str, start_line: int = None, end_line: int = None) -> str:
    """Read file contents with optional line range."""
    if not os.path.exists(path):
        return f"❌ File not found: {path}"
    if os.path.isdir(path):
        return f"❌ This is a directory, not a file: {path}"

    # Check for binary formats that need special handling
    ext = os.path.splitext(path)[1].lower()
    reader = _BINARY_READERS.get(ext)
    if reader:
        try:
            return reader(path)
        except Exception as ex:
            return f"❌ Failed to read {ext} file: {str(ex)}"

    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()

        total = len(lines)
        s = (start_line - 1) if start_line else 0
        e = end_line if end_line else total
        s = max(0, min(s, total))
        e = max(s, min(e, total))

        selected = lines[s:e]

        # Output with line numbers
        numbered = []
        for i, line in enumerate(selected, start=s + 1):
            numbered.append(f"{i:4d} | {line.rstrip()}")

        header = f"📄 {path} ({total} lines total, showing {s+1}-{e})\n"
        return header + "\n".join(numbered)

    except Exception as ex:
        return f"❌ Failed to read file: {str(ex)}"


@tool(
    name="edit_file",
    description="Edit a file using search-and-replace. Performs an exact match on old_text and replaces it with new_text. Prefer this tool for modifying files instead of rewriting the entire file with run_command. A backup of the original file is automatically saved before each edit.",
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "Path to the file to edit.",
            },
            "old_text": {
                "type": "string",
                "description": "The original text to replace. Must exactly match the file content, including whitespace and indentation. Use read_file first to view the file, then copy the section you want to modify.",
            },
            "new_text": {
                "type": "string",
                "description": "The new text to replace the old text with.",
            },
        },
        "required": ["path", "old_text", "new_text"],
    },
)
def edit_file(path: str, old_text: str, new_text: str) -> str:
    """Edit a file using search and replace."""
    if not os.path.exists(path):
        return f"❌ File not found: {path}"

    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()

        count = content.count(old_text)

        if count == 0:
            # Show file snippet to help LLM locate the issue
            lines = content.split("\n")
            preview = "\n".join(lines[:30]) if len(lines) > 30 else content
            return (
                f"❌ No match found. Please verify old_text exactly matches the file content (including whitespace and indentation).\n\n"
                f"First 30 lines:\n{preview}"
            )

        if count > 1:
            # Show all match positions
            positions = []
            start = 0
            for i in range(count):
                pos = content.index(old_text, start)
                line_num = content[:pos].count("\n") + 1
                positions.append(f"  Line {line_num}")
                start = pos + 1
            return (
                f"❌ Found {count} matches, cannot determine which one to replace.\n"
                f"Match positions:\n" + "\n".join(positions) + "\n\n"
                f"Include more surrounding context in old_text to uniquely identify the target."
            )

        # Auto-snapshot before modifying
        _snapshot_file(path)

        # Unique match, perform replacement
        new_content = content.replace(old_text, new_text, 1)
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_content)

        old_lines = len(old_text.splitlines())
        new_lines = len(new_text.splitlines())
        return f"✅ Replaced ({old_lines} lines → {new_lines} lines)"

    except Exception as ex:
        return f"❌ Edit failed: {str(ex)}"


@tool(
    name="write_file",
    description="Create a new file or overwrite an existing file with the given content. Use this to create new files (scripts, configs, documents, etc.). If the file already exists, a backup is automatically saved before overwriting. Parent directories are created automatically if they don't exist. For modifying specific parts of an existing file, prefer edit_file instead.",
    parameters={
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "File path (absolute or relative to workspace). Parent directories will be created if needed.",
            },
            "content": {
                "type": "string",
                "description": "The full content to write to the file.",
            },
            "overwrite": {
                "type": "boolean",
                "description": "Whether to overwrite if the file already exists. Defaults to false. If false and the file exists, the operation will fail with an error.",
                "default": False,
            },
        },
        "required": ["path", "content"],
    },
)
def write_file(path: str, content: str, overwrite: bool = False) -> str:
    """Create a new file or overwrite an existing file."""
    # Track whether file existed before writing
    file_existed = os.path.exists(path)

    # Check if file already exists
    if file_existed:
        if os.path.isdir(path):
            return f"❌ Path is a directory, not a file: {path}"
        if not overwrite:
            return (
                f"❌ File already exists: {path}\n"
                f"Set overwrite=true to replace it, or use edit_file to modify specific parts."
            )
        # Auto-snapshot before overwriting
        _snapshot_file(path)

    try:
        # Auto-create parent directories
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)

        with open(path, "w", encoding="utf-8") as f:
            f.write(content)

        lines = content.count("\n") + (1 if content and not content.endswith("\n") else 0)
        size = len(content.encode("utf-8"))
        size_str = f"{size / 1024:.1f}KB" if size >= 1024 else f"{size}B"
        action = "overwritten" if file_existed else "created"

        return f"✅ File {action}: {path} ({lines} lines, {size_str})"

    except Exception as ex:
        return f"❌ Failed to write file: {str(ex)}"


@tool(
    name="list_snapshots",
    description="List file snapshots created by the Agent before edits. Use this when the user wants to undo an Agent edit or recover a file the Agent accidentally broke. Each snapshot shows the timestamp and original file path.",
    parameters={
        "type": "object",
        "properties": {
            "file_filter": {
                "type": "string",
                "description": "Optional: only show snapshots containing this filename or path fragment.",
                "default": "",
            },
        },
    },
)
def list_snapshots(file_filter: str = "") -> str:
    """List available file snapshots from before Agent edits."""
    try:
        workspace = _get_workspace()
        snap_base = os.path.join(workspace, SNAPSHOT_DIR)

        if not os.path.exists(snap_base):
            return "📂 No snapshots yet. Snapshots are created automatically when edit_file modifies a file."

        entries = []
        for ts_dir in sorted(os.listdir(snap_base), reverse=True):
            ts_path = os.path.join(snap_base, ts_dir)
            if not os.path.isdir(ts_path):
                continue
            for root, _, files in os.walk(ts_path):
                for fname in files:
                    full = os.path.join(root, fname)
                    rel = os.path.relpath(full, ts_path)
                    if file_filter and file_filter not in rel:
                        continue
                    size = os.path.getsize(full)
                    size_str = f"{size / 1024:.1f}KB" if size >= 1024 else f"{size}B"
                    # Parse timestamp from dir name
                    try:
                        ts_display = datetime.strptime(ts_dir, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        ts_display = ts_dir
                    entries.append(f"  📄 {rel}\n     ⏱️ {ts_display} | Size: {size_str}\n     Path: {full}")

        if not entries:
            return "📂 No snapshots found" + (f" matching '{file_filter}'" if file_filter else "") + "."

        result = f"📂 Agent edit snapshots (newest first):\n\n"
        result += "\n".join(entries)
        result += "\n\nTo restore: use run_command(\"cp <snapshot_path> <original_path>\") to copy it back."
        return result

    except Exception as e:
        return f"❌ Failed to list snapshots: {e}"

